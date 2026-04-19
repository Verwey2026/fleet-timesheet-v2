import streamlit as st
import pandas as pd
import re
import io
from datetime import time
from urllib.parse import unquote
import openpyxl

# ===== PASSWORD GATE =====
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("Fleet Timesheet Processor")
    password = st.text_input("Enter password", type="password")
    if password:
        if password == st.secrets["app_password"]:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password")
    st.stop()

# ===== MAIN APP =====
st.set_page_config(page_title="Fleet Timesheet Processor V4.22", layout="wide")
st.title("Fleet Timesheet Processor VERSION 4.22 - Verwey Vervoer")

st.markdown("**Rules:** 195.03 fills M-F first. Sleep out type = LOCAL / XBORDER / NONE from End Location link.")

NORMAL_HOURS_THRESHOLD = 195.03
GEO_FENCE_KEYWORDS = ['MIDDELBURG', 'STEVE TSHWETE']
CROSSBORDER_COUNTRIES = ['ZIMBABWE', 'BOTSWANA', 'NAMIBIA', 'MOZAMBIQUE', 'ZAMBIA', 'DRC', 'LESOTHO', 'ESWATINI', 'MALAWI', 'TANZANIA', 'ANGOLA']
ABNORMAL_FLEETS = ['FL221', 'FL222', 'FL223', 'FL225', 'FL229', 'FL230', 'FL238']

st.sidebar.subheader("NBCRFLI Settings")
NORMAL_HOURS_THRESHOLD = st.sidebar.number_input("Normal Hours Threshold", value=195.03, step=0.01)
st.sidebar.markdown(f"**Abnormal Fleets:** {', '.join(ABNORMAL_FLEETS)}")
st.sidebar.markdown(f"**Geo Fence (0 allowance):** {', '.join(GEO_FENCE_KEYWORDS)}")

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Upload Tracking Report")
    tracking_file = st.file_uploader("Must have End Location column", type=["xlsx", "xls"], key="tracking")

with col2:
    st.subheader("2. Upload Driver Allocation")
    allocation_file = st.file_uploader("Sheet = Driver, Headers: DAY | DATE | FLEET | MEAL HOUR | SLEEP OUT", type=["xlsx", "xls"], key="allocation")

def extract_location_from_cell(cell):
    """Extract location from hyperlink, formula, or value. Handles blank display text."""
    link = ''

    # Priority 1: Excel hyperlink object
    if cell.hyperlink and cell.hyperlink.target:
        link = str(cell.hyperlink.target)
    # Priority 2: HYPERLINK formula =HYPERLINK("url","text")
    elif isinstance(cell.value, str) and cell.value.startswith('=HYPERLINK'):
        match = re.search(r'HYPERLINK\("([^"]+)"', cell.value)
        if match:
            link = match.group(1)
    # Priority 3: Plain text
    else:
        link = str(cell.value) if cell.value else ''

    link = link.upper().strip()
    if not link or link == 'NONE':
        return ''

    # Extract town from Google Maps URL
    if 'GOOGLE.COM/MAPS' in link or 'MAPS.APP.GOO.GL' in link or 'HTTP' in link:
        # Try /place/Town,+Province
        match = re.search(r'/PLACE/([^/@?]+)', link)
        if match:
            return unquote(match.group(1).replace('+', ' '))
        # Try?q=Town
        match = re.search(r'[?&]Q=([^&]+)', link)
        if match:
            return unquote(match.group(1).replace('+', ' '))
        # Try /@lat,lng,zoom
        match = re.search(r'/@[-0-9.]+,[-0-9.]+', link)
        if match:
            return 'COORDINATES_FOUND' # Can't determine town from coords alone

    return link

def read_tracking_with_links(file):
    """Read tracking file. data_only=False to access hyperlinks and formulas."""
    wb = openpyxl.load_workbook(file, data_only=False) # CRITICAL: False to get hyperlinks
    ws = wb.active
    header_row = None
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        row_str = ' '.join([str(x).upper() for x in row if x])
        if 'REGISTRATION' in row_str and 'END LOCATION' in row_str and 'DEPARTURE TIME' in row_str:
            header_row = idx
            break
    if not header_row:
        header_row = 1

    headers = [str(cell.value).strip() if cell.value else f'Unnamed_{i}' for i, cell in enumerate(ws[header_row])]
    end_loc_idx = None
    for i, h in enumerate(headers):
        if 'END LOCATION' in h.upper():
            end_loc_idx = i
            break

    data = []
    for row in ws.iter_rows(min_row=header_row + 1):
        row_data = {}
        for i, cell in enumerate(row):
            col_name = headers[i]
            if i == end_loc_idx:
                row_data[col_name] = extract_location_from_cell(cell)
            else:
                row_data[col_name] = cell.value
        data.append(row_data)
    return pd.DataFrame(data)

def extract_yard_hours_from_text(text):
    if pd.isna(text): return 0.0
    text = str(text).lower()
    match = re.search(r'(\d+\.?\d*)\s*h(?:our)?s?\s*yard|yard[:\s]*(\d+\.?\d*)', text)
    if match: return float(match.group(1) or match.group(2))
    return 0.0

def clean_col_name(col):
    return str(col).strip().lower().replace(':', '').replace('.', '').replace('_', ' ').strip()

def standardize_columns(df):
    rename_map = {
        'date': 'Date', 'trip date': 'Date', 'tripdate': 'Date',
        'driver': 'Employee Name', 'employee': 'Employee Name', 'employee name': 'Employee Name', 'name': 'Employee Name', 'phh': 'Employee Name',
        'notes': 'Activity Description', 'description': 'Activity Description', 'activity': 'Activity Description', 'activity description': 'Activity Description', 'unnamed 11': 'Activity Description',
        'departure time': 'Start Time', 'start time': 'Start Time', 'first movement': 'Start Time', 'departure': 'Start Time', 'starttime': 'Start Time',
        'arrival time': 'End Time', 'end time': 'End Time', 'last movement': 'End Time', 'arrival': 'End Time', 'endtime': 'End Time',
        'fleet': 'Fleet Number', 'vehicle': 'Fleet Number', 'truck': 'Fleet Number', 'fleet no': 'Fleet Number',
        'reg': 'Fleet Number', 'registration': 'Fleet Number', 'registration nr': 'Fleet Number', 'reg nr': 'Fleet Number', 'fleet number': 'Fleet Number', 'reg no': 'Fleet Number', 'registration nr.': 'Fleet Number',
        'meal hour': 'Meal Hour', 'meal': 'Meal Hour', 'meal hours': 'Meal Hour',
        'sleep out': 'Sleep Out', 'sleep': 'Sleep Out', 'sleepout': 'Sleep Out',
        'end location': 'End Location', 'arrival location': 'End Location', 'last location': 'End Location', 'destination': 'End Location',
        'start location': 'Start Location', 'departure location': 'Start Location', 'origin': 'Start Location'
    }
    df.columns = [clean_col_name(col) for col in df.columns]
    for old, new in rename_map.items():
        df.columns = [new if old == col else col for col in df.columns]
    df = df.loc[:, ~df.columns.duplicated()]
    return df

def auto_lunch_deduction(row):
    meal = pd.to_numeric(row.get('Meal Hour', 0), errors='coerce')
    if pd.notna(meal) and meal > 0:
        return meal
    start = row['Start Time']
    end = row['End Time']
    gross = (end - start).total_seconds() / 3600
    if gross > 5 and start.time() <= time(12, 0) and end.time() >= time(13, 0):
        return 1.0
    return 0.0

def classify_sleep_out_type(row):
    """Returns: 'LOCAL', 'XBORDER', 'NONE', or 'UNKNOWN'"""
    nights = pd.to_numeric(row.get('Sleep Out', 0), errors='coerce')
    if pd.isna(nights) or nights == 0:
        return 'NONE'

    location = str(row.get('End Location', '')).upper().strip()

    if not location or location == 'COORDINATES_FOUND':
        return 'UNKNOWN' # Sleep out claimed but location can't be determined

    if any(country in location for country in CROSSBORDER_COUNTRIES):
        return 'XBORDER'
    if any(town in location for town in GEO_FENCE_KEYWORDS):
        return 'NONE' # Geo fence = 0 allowance
    return 'LOCAL'

def is_abnormal(fleet_no):
    return any(fleet_no.startswith(prefix) for prefix in ABNORMAL_FLEETS)

if tracking_file and allocation_file:
    try:
        df_track = read_tracking_with_links(tracking_file)
        df_track = standardize_columns(df_track)

        xls = pd.ExcelFile(allocation_file)
        all_alloc_dfs = []
        for sheet_name in xls.sheet_names:
            df_sheet = pd.read_excel(xls, sheet_name=sheet_name)
            df_sheet = standardize_columns(df_sheet)
            df_sheet['Employee Name'] = sheet_name.strip()
            all_alloc_dfs.append(df_sheet)

        if not all_alloc_dfs:
            st.error("No valid sheets found in allocation file")
            st.stop()

        df_alloc = pd.concat(all_alloc_dfs, ignore_index=True)

        with st.expander("Debug: Allocation file check"):
            st.write(f"Allocation columns: {list(df_alloc.columns)}")
            if 'Sleep Out' in df_alloc.columns:
                st.write(f"Sleep Out total nights: {df_alloc['Sleep Out'].sum():.0f}")
            else:
                st.error("SLEEP OUT column not found! Check header spelling.")
            st.dataframe(df_alloc.head())

        # Date handling
        df_track['Date'] = pd.to_datetime(df_track['Start Time'], errors='coerce', dayfirst=True)
        df_alloc['Date'] = pd.to_datetime(df_alloc['Date'], format='%Y %m %d', errors='coerce')
        df_track['Date'] = df_track['Date'].dt.strftime('%Y-%m-%d')
        df_alloc['Date'] = df_alloc['Date'].dt.strftime('%Y-%m-%d')

        # Standardize keys
        df_track['Fleet Number'] = df_track['Fleet Number'].astype(str).str.strip().str.upper()
        df_alloc['Fleet Number'] = df_alloc['Fleet Number'].astype(str).str.strip().str.upper()

        df_track = df_track.dropna(subset=['Fleet Number', 'Date', 'Start Time', 'End Time'])
        df_alloc = df_alloc.dropna(subset=['Fleet Number', 'Date'])

        merge_cols = ['Fleet Number', 'Date', 'Employee Name']
        if 'Sleep Out' in df_alloc.columns:
            merge_cols.append('Sleep Out')
        if 'Meal Hour' in df_alloc.columns:
            merge_cols.append('Meal Hour')

        df_merged = pd.merge(df_track, df_alloc[merge_cols], on=['Fleet Number', 'Date'], how='left', indicator='MERGE_CHECK')

        df_merged['Employee Name'] = df_merged['Employee Name'].fillna('UNALLOCATED')
        df_merged['Sleep Out'] = pd.to_numeric(df_merged.get('Sleep Out', 0), errors='coerce').fillna(0)
        df_merged['Meal Hour'] = pd.to_numeric(df_merged.get('Meal Hour', 0), errors='coerce').fillna(0)

        if df_merged.empty:
            st.error("No data after processing.")
            st.stop()

        # ===== HOURS CALCULATION =====
        df_merged['Start Time'] = pd.to_datetime(df_merged['Start Time'], errors='coerce')
        df_merged['End Time'] = pd.to_datetime(df_merged['End Time'], errors='coerce')
        df_merged['gross_hours'] = (df_merged['End Time'] - df_merged['Start Time']).dt.total_seconds() / 3600

        df_merged['meal_hour'] = df_merged.apply(auto_lunch_deduction, axis=1)

        # Sleep out classification
        df_merged['SLEEP OUT RAW'] = df_merged['Sleep Out']
        df_merged['SLEEP OUT TYPE'] = df_merged.apply(classify_sleep_out_type, axis=1)
        df_merged['sleep_out_local'] = df_merged.apply(lambda r: r['Sleep Out'] if r['SLEEP OUT TYPE'] == 'LOCAL' else 0, axis=1)
        df_merged['sleep_out_crossborder'] = df_merged.apply(lambda r: r['Sleep Out'] if r['SLEEP OUT TYPE'] == 'XBORDER' else 0, axis=1)

        # Yard hours
        if 'Activity Description' in df_merged.columns:
            df_merged['yard_hours'] = df_merged['Activity Description'].apply(lambda x: extract_yard_hours_from_text(str(x)))
        else:
            df_merged['yard_hours'] = 0.0

        df_merged['total_hours'] = (df_merged['gross_hours'] - df_merged['meal_hour']).clip(lower=0)
        df_merged['driving_hours'] = (df_merged['total_hours'] - df_merged['yard_hours']).clip(lower=0)

        df_merged['weekday'] = pd.to_datetime(df_merged['Date']).dt.dayofweek
        df_merged['is_abnormal'] = df_merged['Fleet Number'].apply(is_abnormal)

        df_merged = df_merged.sort_values(['Employee Name', 'Date', 'Start Time'])

        # ===== 195.03 LOGIC =====
        df_merged['normal_weekday'] = 0.0
        df_merged['normal_sat'] = 0.0
        df_merged['normal_sun'] = 0.0
        df_merged['ot_weekday'] = 0.0
        df_merged['ot_sat'] = 0.0
        df_merged['ot_sun'] = 0.0

        for driver in df_merged['Employee Name'].unique():
            if driver == 'UNALLOCATED':
                continue
            driver_mask = df_merged['Employee Name'] == driver

            # M-F cap
            mf_mask = driver_mask & (df_merged['weekday'] < 5)
            mf_cumsum = 0.0
            for idx in df_merged[mf_mask].index:
                hours_today = df_merged.at[idx, 'total_hours']
                normal_today = min(hours_today, max(0, NORMAL_HOURS_THRESHOLD - mf_cumsum))
                ot_today = hours_today - normal_today
                df_merged.at[idx, 'normal_weekday'] = normal_today
                df_merged.at[idx, 'ot_weekday'] = ot_today
                mf_cumsum += normal_today

            # Sat
            normal_remaining = max(0, NORMAL_HOURS_THRESHOLD - mf_cumsum)
            sat_mask = driver_mask & (df_merged['weekday'] == 5)
            for idx in df_merged[sat_mask].index:
                hours_today = df_merged.at[idx, 'total_hours']
                normal_today = min(hours_today, normal_remaining)
                ot_today = hours_today - normal_today
                df_merged.at[idx, 'normal_sat'] = normal_today
                df_merged.at[idx, 'ot_sat'] = ot_today
                normal_remaining -= normal_today

            # Sun
            sun_mask = driver_mask & (df_merged['weekday'] == 6)
            for idx in df_merged[sun_mask].index:
                hours_today = df_merged.at[idx, 'total_hours']
                normal_today = min(hours_today, normal_remaining)
                ot_today = hours_today - normal_today
                df_merged.at[idx, 'normal_sun'] = normal_today
                df_merged.at[idx, 'ot_sun'] = ot_today
                normal_remaining -= normal_today

        df_merged['total_ot_15'] = df_merged['ot_weekday'] + df_merged['ot_sat']

        # ===== SUMMARY =====
        abnormal_days = df_merged[df_merged['is_abnormal'] == True].groupby('Employee Name')['Date'].nunique().reset_index()
        abnormal_days = abnormal_days.rename(columns={'Date': 'ABNORMAL DAYS'})

        driver_totals = df_merged.groupby('Employee Name').agg({
            'total_hours': 'sum',
            'normal_weekday': 'sum',
            'normal_sat': 'sum',
            'normal_sun': 'sum',
            'ot_weekday': 'sum',
            'ot_sat': 'sum',
            'ot_sun': 'sum',
            'total_ot_15': 'sum',
            'yard_hours': 'sum',
            'driving_hours': 'sum',
            'meal_hour': 'sum',
            'sleep_out_local': 'sum',
            'sleep_out_crossborder': 'sum'
        }).reset_index()

        driver_totals = pd.merge(driver_totals, abnormal_days, on='Employee Name', how='left')
        driver_totals['ABNORMAL DAYS'] = driver_totals['ABNORMAL DAYS'].fillna(0).astype(int)

        driver_totals = driver_totals.rename(columns={
            'total_hours': 'PAID HOURS',
            'normal_weekday': 'NORMAL M-F',
            'normal_sat': 'NORMAL SAT@1.5',
            'normal_sun': 'NORMAL SUN@2.0',
            'ot_weekday': 'OT M-F@1.5',
            'ot_sat': 'OT SAT@1.5',
            'ot_sun': 'SUN@2.0',
            'total_ot_15': 'TOTAL OT @1.5',
            'yard_hours': 'YARD',
            'driving_hours': 'DRIVING',
            'meal_hour': 'UNPAID MEAL',
            'sleep_out_local': 'SLEEP OUT LOCAL',
            'sleep_out_crossborder': 'SLEEP OUT XBORDER'
        })

        driver_totals = driver_totals[driver_totals['Employee Name']!= 'UNALLOCATED']

        st.success(f"Allocated {len(df_merged)} trips to {df_merged['Employee Name'].nunique()} drivers!")

        unallocated = df_merged[df_merged['Employee Name'] == 'UNALLOCATED']
        if len(unallocated) > 0:
            st.warning(f"{len(unallocated)} trips unallocated. Check Date + Fleet in allocation file.")

        unknown_sleep = len(df_merged[df_merged['SLEEP OUT TYPE'] == 'UNKNOWN'])
        if unknown_sleep > 0:
            st.warning(f"{unknown_sleep} trips have SLEEP OUT > 0 but End Location couldn't be read from link. Check tracking file.")

        total_sleep_raw = df_merged['SLEEP OUT RAW'].sum()
        total_sleep_local = df_merged['sleep_out_local'].sum()
        total_sleep_xborder = df_merged['sleep_out_crossborder'].sum()
        st.info(f"Sleep Out Totals: RAW={total_sleep_raw:.0f} | LOCAL={total_sleep_local:.0f} | XBORDER={total_sleep_xborder:.0f} | UNKNOWN={unknown_sleep}")

        st.subheader(f"Summary - M-F caps at {NORMAL_HOURS_THRESHOLD}h")
        st.dataframe(driver_totals.round(2))

        st.subheader("All Trips - Check SLEEP OUT TYPE + End Location")
        display_cols = ['Date', 'Employee Name', 'Fleet Number', 'End Location', 'SLEEP OUT RAW', 'SLEEP OUT TYPE', 'weekday',
                       'total_hours', 'normal_weekday', 'normal_sat', 'normal_sun',
                       'ot_weekday', 'ot_sat', 'ot_sun', 'total_ot_15',
                       'sleep_out_local', 'sleep_out_crossborder']
        display_cols = [col for col in display_cols if col in df_merged.columns]
        st.dataframe(df_merged[display_cols].round(2))

        # ===== EXCEL EXPORT =====
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            driver_totals.to_excel(writer, index=False, sheet_name='SUMMARY')
            df_merged[display_cols].to_excel(writer, index=False, sheet_name='ALL TRIPS')

            for driver in sorted(df_merged['Employee Name'].unique()):
                if driver == 'UNALLOCATED':
                    continue
                df_driver = df_merged[df_merged['Employee Name'] == driver][display_cols].copy()
                subtotal_data = {col: [''] for col in display_cols}
                subtotal_data['Date'] = ['TOTAL']
                subtotal_data['Employee Name'] = [driver]
                for col in ['total_hours', 'normal_weekday', 'normal_sat', 'normal_sun',
                           'ot_weekday', 'ot_sat', 'ot_sun', 'total_ot_15', 'sleep_out_local', 'sleep_out_crossborder']:
                    if col in df_driver.columns:
                        subtotal_data[col] = [df_driver[col].sum()]
                subtotal = pd.DataFrame(subtotal_data)
                df_driver_out = pd.concat([df_driver, subtotal], ignore_index=True)
                sheet_name = re.sub(r'[\\/*?:\[\]]', '', driver)[:31]
                df_driver_out.to_excel(writer, index=False, sheet_name=sheet_name)

        output.seek(0)

        st.download_button(
            "📥 Download Excel - V4.22",
            output,
            "fleet_timesheet_processed.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error: {str(e)}")
        st.exception(e)
else:
    st.info("👆 Upload both files to start processing")
